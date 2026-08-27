from __future__ import annotations

import base64
import json
import os
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import fitz
import pytest
import balance_anonymizer.adapters.pdf as pdf_adapter
import yaml
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

from anonymize_balances import main
from balance_anonymizer.adapters import XlsxAdapter, XmlAdapter
from balance_anonymizer.adapters.base import AdapterError
from balance_anonymizer.batch import (
    BatchProcessor,
    _blocking_conflicts,
    list_input_files,
    resolve_groups,
)
from balance_anonymizer.manifest import ManifestGroup, load_manifest
from balance_anonymizer.models import (
    Category,
    DocumentSnapshot,
    FormatLocation,
    LedgerLine,
    OwnerIdentity,
    RelationType,
    TemporalMetadata,
)
from balance_anonymizer.pseudonyms import Pseudonymizer, canonical_association_key
from balance_anonymizer.registry import PseudonymRegistry
from balance_anonymizer.relations import (
    _comparison_maps,
    infer_pair,
    infer_relations,
    normalize_account_code,
)
from .conftest import AccountingRow, CLASSIC_OWNER, CLASSIC_RFC, make_classic_balance


SEED = "semilla-multiformato-local-2026"
XML_NS = "http://www.sat.gob.mx/esquemas/ContabilidadE/1_3/BalanzaComprobacion"
XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
PNG_1X1 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def _line(
    code: str,
    values: tuple[str, str, str, str] = ("100.00", "2.00", "1.00", "101.00"),
) -> LedgerLine:
    names = ("saldo_inicial", "debe", "haber", "saldo_final")
    return LedgerLine(
        code,
        normalize_account_code(code),
        None,
        "CUENTA ORDINARIA",
        dict(zip(names, map(Decimal, values))),
        dict(zip(names, values)),
        FormatLocation("test"),
    )


def _snapshot(
    path: Path,
    lines: list[LedgerLine],
    *,
    adapter: str = "test",
    owner: str | None = "ENTIDAD DE PRUEBA",
    rfc: str | None = "EDP240101AA1",
    year: int = 2024,
    month: int = 5,
) -> DocumentSnapshot:
    return DocumentSnapshot(
        path,
        adapter,
        "TEST",
        OwnerIdentity(owner, rfc),
        TemporalMetadata(year=year, month=month),
        lines,
    )


def _make_xml(
    path: Path,
    *,
    rows: tuple[tuple[str, str, str, str, str], ...] = (
        ("101.01", "1000.00", "200.00", "50.00", "1150.00"),
    ),
    rfc: str = CLASSIC_RFC,
    year: int = 2025,
    month: int = 2,
    fecha_mod: str | None = None,
    signed: bool = False,
) -> None:
    signature = (
        ' Sello="SELLO_DE_PRUEBA" Certificado="CERTIFICADO_DE_PRUEBA" noCertificado="123456"'
        if signed
        else ""
    )
    modified = f' FechaModBal="{fecha_mod}"' if fecha_mod else ""
    accounts = "".join(
        f'<BCE:Ctas NumCta="{code}" SaldoIni="{initial}" Debe="{debit}" '
        f'Haber="{credit}" SaldoFin="{final}"/>'
        for code, initial, debit, credit, final in rows
    )
    path.write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        f'<BCE:Balanza xmlns:BCE="{XML_NS}" xmlns:xsi="{XSI_NS}" '
        f'xsi:schemaLocation="{XML_NS} local-bal.xsd" Version="1.3" RFC="{rfc}" '
        f'Anio="{year}" Mes="{month:02d}" TipoEnvio="N"{modified}{signature}>'
        f"{accounts}</BCE:Balanza>",
        encoding="utf-8",
    )


def _make_xlsx_b(
    path: Path,
    *,
    owner: str = CLASSIC_OWNER,
    rfc: str = CLASSIC_RFC,
    rows: tuple[tuple[str, str, object, object, object, object], ...] = (
        ("101.01", "CAJA GENERAL", "1000.00", "200.00", "50.00", "1150.00"),
    ),
    year: int = 2025,
    month: int = 2,
) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "Balanza"
    sheet["A1"] = f"{owner} {rfc}"
    sheet["A2"] = f"Periodo: {year}-{month:02d}"
    for column, value in enumerate(
        ("Cuenta", "Saldo inicial", "Debe", "Haber", "Saldo final"), start=1
    ):
        sheet.cell(4, column, value)
    for row_number, row in enumerate(rows, start=5):
        code, description, *amounts = row
        sheet.cell(row_number, 1, f"{code} {description}")
        for column, value in enumerate(amounts, start=2):
            sheet.cell(row_number, column, value)
    book.save(path)


def _make_xlsx_a(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "Balanza"
    sheet["A1"] = "ENTIDAD PROPIETARIA DEMOSTRACION"
    sheet["H1"] = "CONTPAQ i"
    sheet["A3"] = "Balanza al 31/05/2024"
    sheet["A4"] = "Fecha de impresión: 08/06/2024"
    sheet["A6"], sheet["B6"] = "Cuenta", "Nombre"
    sheet["C6"], sheet["E6"], sheet["F6"], sheet["G6"] = (
        "Saldo inicial", "Cargos", "Abonos", "Saldo final"
    )
    sheet["C7"], sheet["D7"], sheet["G7"], sheet["H7"] = (
        "Deudor", "Acreedor", "Deudor", "Acreedor"
    )
    descriptions = (
        "BANCO DEMOSTRACION CTA. 1234",
        "PERSONA DEMOSTRACION CTA. 9876",
        "BANCO SIN IDENTIFICADOR",
    )
    for index, description in enumerate(descriptions, start=8):
        sheet.cell(index, 1, f"10{index}.01")
        sheet.cell(index, 2, description)
        for column, value in enumerate((100, 0, 20, 5, 115, 0), start=3):
            sheet.cell(index, column, value)
    book.save(path)


def _make_xlsx_c(path: Path, *, include_logo: bool = True) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "Balanza principal"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "ENTIDAD CON IMAGEN DE PRUEBA"
    sheet["A2"] = "RFC: ECI240101AA1"
    sheet["A3"] = "Moneda: MXN"
    sheet["A4"] = "Periodo: 2024-05"
    sheet["A5"] = "Fecha de impresión: 07/06/2024"
    sheet["D1"] = "Dirección: VIA ORIGINAL #123"
    sheet["D2"] = "Población: CIUDAD ORIGINAL"
    sheet["D3"] = "Cédula: CED-1234-AA"
    sheet.oddFooter.center.text = "ENTIDAD CON IMAGEN DE PRUEBA - 07/06/2024"
    headers = ("No. Cuenta", "Naturaleza", "Cuenta", "Saldo inicial", "Debe", "Haber", "Saldo final")
    for column, value in enumerate(headers, start=1):
        sheet.cell(7, column, value)
    rows = (
        ("101.01", "D", "PERSONA SIN IDENTIFICADOR", 100, 2, 1, 101),
        ("102.01", "D", "BANCO SIN IDENTIFICADOR", 200, 3, 1, 202),
        ("103.01", "D", "EMPRESA SIN IDENTIFICADOR", 300, 4, 2, 302),
    )
    for row_number, row in enumerate(rows, start=8):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row_number, column, value)
            if row_number == 8 and column == 4:
                cell.number_format = '#,##0.00'
                cell.font = Font(name="Arial", bold=True, color="FF112233")
                cell.fill = PatternFill("solid", fgColor="FFFFFF00")
    sheet["H8"] = "=D8+E8-F8"
    hidden = book.create_sheet("Oculta")
    hidden.sheet_state = "veryHidden"
    hidden["A1"] = "ENTIDAD CON IMAGEN DE PRUEBA"
    hidden["B1"].comment = Comment("RFC: ECI240101AA1", "Autor")
    book.properties.creator = "ENTIDAD CON IMAGEN DE PRUEBA"
    book.properties.created = datetime(2024, 6, 7, 12, 30, 0)
    if include_logo:
        png = path.with_suffix(".png")
        png.write_bytes(PNG_1X1)
        logo = Image(png)
        logo.width, logo.height = 120, 45
        logo.anchor = "E1"
        sheet.add_image(logo)
    book.save(path)


def _process(sources: list[Path], output: Path, manifest: list[ManifestGroup] | None = None):
    with PseudonymRegistry(output / "registry.sqlite") as registry:
        return BatchProcessor(SEED, registry=registry).run(
            [item.resolve() for item in sources], output, manifest=manifest or []
        )


def test_relations_exact_projection_series_and_collision_are_conservative(tmp_path: Path) -> None:
    exact_left = _snapshot(
        tmp_path / "a", [_line("001-02"), _line("003-04"), _line("007-08")]
    )
    exact_right = _snapshot(
        tmp_path / "b", [_line("001.02"), _line("003.04"), _line("007.08")]
    )
    exact = infer_pair(exact_left, exact_right)
    assert exact and exact.relation == RelationType.EXACT_EQUIVALENT

    projection = infer_pair(
        exact_left,
        _snapshot(
            tmp_path / "c",
            [_line("001.02"), _line("003.04"), _line("007.08"), _line("005.06")],
        ),
    )
    assert projection and projection.relation == RelationType.PROJECTION

    earlier = _snapshot(
        tmp_path / "d", [_line("100", ("0", "10", "2", "8"))], year=2024, month=5
    )
    later = _snapshot(
        tmp_path / "e", [_line("100", ("8", "3", "1", "10"))], year=2024, month=6
    )
    # La regla de serie exige al menos tres renglones para evitar falsos positivos.
    earlier.ledger_lines *= 3
    later.ledger_lines = [
        _line(str(100 + index), ("8", "3", "1", "10")) for index in range(3)
    ]
    earlier.ledger_lines = [
        _line(str(100 + index), ("0", "10", "2", "8")) for index in range(3)
    ]
    series = infer_pair(earlier, later)
    assert series and series.relation == RelationType.SERIES

    collision = _snapshot(tmp_path / "f", [_line("001-02"), _line("001.02")])
    ambiguous = infer_pair(collision, exact_right)
    assert ambiguous and ambiguous.relation == RelationType.AMBIGUOUS
    assert ambiguous.conflicts == ["ACCOUNT_NORMALIZATION_COLLISION"]


def test_registry_is_persistent_namespaced_and_contains_no_originals(tmp_path: Path) -> None:
    registry_path = tmp_path / "anon.sqlite"
    original = "IDENTIFICADOR PRIVADO 445566"
    with PseudonymRegistry(registry_path) as registry:
        first = Pseudonymizer(SEED, registry=registry, scope="entity_001")
        assigned = {
            first.compact_company(original),
            first.rfc("ABC240101AA1", original),
            first.bank(original),
            first.numeric_identifier("4455-66", original),
            first.address(original),
            str(first.month_offset()),
        }
    with PseudonymRegistry(registry_path) as registry:
        second = Pseudonymizer(SEED, registry=registry, scope="entity_001")
        assert second.compact_company(original) in assigned
        assert second.numeric_identifier("4455-66", original) in assigned
        rows = registry._connection.execute(
            "SELECT namespace, identifier_hmac, synthetic_value, algorithm_version FROM assignments"
        ).fetchall()
    assert {row[0] for row in rows} >= {"owner", "rfc", "bank", "account", "address", "date"}
    assert all(re.fullmatch(r"[A-F0-9]{64}", row[1]) for row in rows)
    raw = registry_path.read_bytes()
    assert original.encode() not in raw and SEED.encode() not in raw


def test_associated_identifier_keeps_punctuation_with_shared_digits() -> None:
    pseudo = Pseudonymizer(SEED)
    key_with_prefix = canonical_association_key("BANCO DEMO CTA.", "12-34")
    key_without_prefix = canonical_association_key("BANCO DEMO", "1234")
    assert key_with_prefix == key_without_prefix
    separated = pseudo.numeric_identifier("12-34", key_with_prefix)
    compact = pseudo.numeric_identifier("1234", key_without_prefix)
    assert separated[2] == "-"
    assert separated.replace("-", "") == compact


def test_pdf_output_uses_positioned_fallback_only_for_exact_expected_codes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    expected = [_line("101.01"), _line("102.01")]
    incomplete = [_line("102.01")]
    monkeypatch.setattr(pdf_adapter, "_web_ledger", lambda document: incomplete)
    monkeypatch.setattr(
        pdf_adapter,
        "_positioned_ledger",
        lambda document, expected_codes=None: expected,
    )

    selected = pdf_adapter._generated_ledger(
        object(),
        "WEB_BALANCE",
        expected,
    )

    assert selected is expected


def test_shared_plan_is_applied_to_pdf_xlsx_and_xml(tmp_path: Path) -> None:
    pdf = tmp_path / "one.pdf"
    xlsx = tmp_path / "two.xlsx"
    xml = tmp_path / "three.xml"
    make_classic_balance(
        pdf,
        pages=1,
        rows_by_page=((AccountingRow("101.01", "CAJA GENERAL"),),),
    )
    _make_xlsx_b(xlsx)
    _make_xml(xml)
    files = (pdf.resolve(), xlsx.resolve(), xml.resolve())
    manifest = [
        ManifestGroup("equivalent_001", RelationType.EXACT_EQUIVALENT, files, "entity_001", xlsx.resolve())
    ]

    run = _process([pdf, xlsx, xml], tmp_path / "out", manifest)

    assert len(run.results) == 3 and all(item.success for item in run.results)
    assert {item.atomic_state for item in run.results} == {"GROUP_COMMITTED"}
    outputs = {item.adapter: Path(item.output or "") for item in run.results}
    with fitz.open(outputs["pdf"]) as document:
        pdf_text = "\n".join(page.get_text("text") for page in document)
    workbook = load_workbook(outputs["xlsx"], data_only=False)
    xlsx_owner = str(workbook.active["A1"].value)
    xml_root = etree.parse(outputs["xml"]).getroot()
    xml_rfc = xml_root.attrib["RFC"]
    synthetic_rfc = re.search(r"Z{3,4}\d{6}[A-Z0-9]{3}", xlsx_owner)
    assert synthetic_rfc
    assert synthetic_rfc.group(0) == xml_rfc and xml_rfc in pdf_text
    synthetic_owner = xlsx_owner.replace(xml_rfc, "").strip()
    assert synthetic_owner in pdf_text


def test_xlsx_owner_without_rfc_text_logo_and_associated_identifiers(tmp_path: Path) -> None:
    source = tmp_path / "standalone.xlsx"
    _make_xlsx_a(source)
    source_digest = source.read_bytes()

    run = _process([source], tmp_path / "out")

    assert len(run.results) == 1 and run.results[0].success
    assert run.results[0].relation == RelationType.STANDALONE.value
    output = Path(run.results[0].output or "")
    book = load_workbook(output, data_only=False)
    values = [str(cell.value or "") for row in book.active.iter_rows() for cell in row]
    assert all("CONTPAQ" not in value.upper() for value in values)
    assert sum("CTA." in value for value in values) == 2
    assert "BANCO SIN IDENTIFICADOR" in values
    assert source.read_bytes() == source_digest
    assert run.results[0].redactions[Category.ASSOCIATED_BANK.value] == 1
    assert run.results[0].redactions[Category.ASSOCIATED_ENTITY.value] == 1


def test_xlsx_removes_logo_relationships_and_preserves_structure_formula_types_styles(tmp_path: Path) -> None:
    source = tmp_path / "merged.xlsx"
    _make_xlsx_c(source)
    original = load_workbook(source, data_only=False)
    original_style = original["Balanza principal"]["D8"]._style

    run = _process([source], tmp_path / "out")

    assert run.results[0].success
    output = Path(run.results[0].output or "")
    generated = load_workbook(output, data_only=False)
    sheet = generated["Balanza principal"]
    assert tuple(map(str, sheet.merged_cells.ranges)) == ("A1:C1",)
    assert sheet["H8"].value == "=D8+E8-F8" and sheet["H8"].data_type == "f"
    assert sheet["D8"]._style == original_style and isinstance(sheet["D8"].value, int)
    assert generated["Oculta"].sheet_state == "veryHidden"
    assert len(sheet._images) == 0
    descriptions = {sheet.cell(row, 3).value for row in range(8, 11)}
    assert descriptions == {
        "PERSONA SIN IDENTIFICADOR",
        "BANCO SIN IDENTIFICADOR",
        "EMPRESA SIN IDENTIFICADOR",
    }
    all_text = "\n".join(
        str(cell.value or "") for item in generated.worksheets for row in item.iter_rows() for cell in row
    )
    assert "VIA ORIGINAL" not in all_text
    assert "CIUDAD ORIGINAL" not in all_text
    assert "CED-1234-AA" not in all_text
    assert "ENTIDAD CON IMAGEN DE PRUEBA" not in (sheet.oddFooter.center.text or "")
    assert "07/06/2024" not in (sheet.oddFooter.center.text or "")
    assert generated.properties.created.date() != date(2024, 6, 7)
    with zipfile.ZipFile(output) as archive:
        names = archive.namelist()
        assert not any(name.startswith("xl/media/") for name in names)
        assert not any(name.lower().startswith("xl/drawings/drawing") for name in names)


def test_xlsx_preserves_ooxml_numeric_text_and_tolerates_custom_sheet_width(tmp_path: Path) -> None:
    source = tmp_path / "producer_variant.xlsx"
    _make_xlsx_b(
        source,
        rows=(("101.01", "CAJA GENERAL", 0.1, 2.0, 1.0, 1.1),),
    )
    book = load_workbook(source, data_only=False)
    book.active["A4"] = "No. de Cuenta Descripción"
    book.save(source)

    with zipfile.ZipFile(source) as archive:
        infos = archive.infolist()
        parts = {item.filename: archive.read(item.filename) for item in infos}
    sheet = etree.fromstring(parts["xl/worksheets/sheet1.xml"])
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sheet.find(f"{{{namespace}}}sheetFormatPr").set("customWidth", "1")
    amount = sheet.find(f".//{{{namespace}}}c[@r='B5']/{{{namespace}}}v")
    assert amount is not None
    amount.text = "0.10000000000000001"
    parts["xl/worksheets/sheet1.xml"] = etree.tostring(
        sheet, encoding="UTF-8", xml_declaration=True
    )
    rewritten = tmp_path / "rewritten.xlsx"
    with zipfile.ZipFile(rewritten, "w") as archive:
        for info in infos:
            archive.writestr(info, parts[info.filename])
    rewritten.replace(source)

    run = _process([source], tmp_path / "out")

    assert run.results[0].success
    output = Path(run.results[0].output or "")
    with zipfile.ZipFile(output) as archive:
        generated = etree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    assert "customWidth" not in generated.find(f"{{{namespace}}}sheetFormatPr").attrib
    generated_amount = generated.find(f".//{{{namespace}}}c[@r='B5']/{{{namespace}}}v")
    assert generated_amount is not None and generated_amount.text == "0.10000000000000001"
    load_workbook(output, data_only=False)


def test_legacy_xls_is_reported_explicitly_instead_of_ignored(tmp_path: Path) -> None:
    source = tmp_path / "legacy.xls"
    source.write_bytes(bytes.fromhex("D0CF11E0A1B11AE1") + b"legacy-test")

    sources = list_input_files(tmp_path)
    run = BatchProcessor(SEED, registry=None).run(
        sources,
        tmp_path / "out",
        discover_only=True,
    )

    assert sources == [source.resolve()]
    assert len(run.results) == 1
    assert run.results[0].adapter == "xls"
    assert run.results[0].extra["codigo_error"] == "XLSX_PROFILE_UNRECOGNIZED"


def test_batch_preserves_safe_pdf_and_xlsx_discovery_codes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pdf = tmp_path / "unsupported.pdf"
    xlsx = tmp_path / "unsupported.xlsx"
    pdf.write_bytes(b"%PDF-test")
    xlsx.write_bytes(b"not-an-ooxml-container")
    processor = BatchProcessor(SEED, registry=None)

    class FailingPdf:
        name = "pdf"

        def discover(self, *args: object, **kwargs: object) -> object:
            raise AdapterError("PDF_PROFILE_UNRECOGNIZED")

    monkeypatch.setitem(processor.adapters, ".pdf", FailingPdf())
    run = processor.run([pdf.resolve(), xlsx.resolve()], tmp_path / "out", discover_only=True)

    codes = {item.adapter: item.extra["codigo_error"] for item in run.results}
    assert codes == {"pdf": "PDF_PROFILE_UNRECOGNIZED", "xlsx": "XLSX_INVALID_CONTAINER"}


def test_xml_preserves_namespace_order_and_decimal_text_while_anonymizing_temporal_fields(tmp_path: Path) -> None:
    source = tmp_path / "balance.xml"
    rows = (
        ("001-02", "-0001.2300", "000.00", "5.500", "-0006.7300"),
        ("003.004", "0", "2.00", "1.00", "1.00"),
    )
    _make_xml(source, rows=rows, fecha_mod="28/02/2025")
    before = XmlAdapter().discover(source, Pseudonymizer(SEED), strict=True)

    run = _process([source], tmp_path / "out")

    assert run.results[0].success
    output = Path(run.results[0].output or "")
    after = XmlAdapter().discover(output, Pseudonymizer(SEED), strict=True)
    assert after.structural["prefix"] == before.structural["prefix"] == "BCE"
    assert after.structural["namespace"] == before.structural["namespace"]
    assert after.structural["schema_location"] == before.structural["schema_location"]
    assert [line.account_code for line in after.ledger_lines] == [line.account_code for line in before.ledger_lines]
    assert [line.amount_representations for line in after.ledger_lines] == [
        line.amount_representations for line in before.ledger_lines
    ]
    root = etree.parse(output).getroot()
    assert root.attrib["RFC"] != CLASSIC_RFC
    assert (root.attrib["Anio"], root.attrib["Mes"], root.attrib["FechaModBal"]) != (
        "2025", "02", "28/02/2025"
    )


def test_xml_rejects_dtd_and_requires_explicit_signature_stripping(tmp_path: Path) -> None:
    dtd = tmp_path / "dtd.xml"
    dtd.write_text(
        '<?xml version="1.0"?><!DOCTYPE Balanza [<!ENTITY xxe SYSTEM "file:///etc/passwd">]>'
        f'<Balanza xmlns="{XML_NS}" Version="1.3" RFC="{CLASSIC_RFC}" Anio="2025" Mes="02" TipoEnvio="N">'
        '<Ctas NumCta="1" SaldoIni="0" Debe="0" Haber="0" SaldoFin="0"/></Balanza>',
        encoding="utf-8",
    )
    with pytest.raises(AdapterError, match="XML_DTD_OR_ENTITY_FORBIDDEN"):
        XmlAdapter().discover(dtd, Pseudonymizer(SEED), strict=True)

    signed = tmp_path / "signed.xml"
    _make_xml(signed, signed=True)
    with pytest.raises(AdapterError, match="SIGNATURE_PRESENT"):
        XmlAdapter().discover(signed, Pseudonymizer(SEED), strict=True)

    rejected = BatchProcessor(SEED, registry=None).run(
        [signed.resolve()], tmp_path / "rejected", discover_only=True
    )
    assert rejected.results[0].extra["codigo_error"] == "SIGNATURE_PRESENT"

    with PseudonymRegistry(tmp_path / "out" / "registry.sqlite") as registry:
        run = BatchProcessor(SEED, registry=registry, strip_signature=True).run(
            [signed.resolve()], tmp_path / "out"
        )
    assert run.results[0].success
    output = Path(run.results[0].output or "")
    raw = output.read_bytes()
    assert b"Sello=" not in raw and b"Certificado=" not in raw and b"noCertificado=" not in raw
    assert b"anonymized-not-for-fiscal-submission" in raw


def test_strict_conflict_needs_manifest_metadata_source(tmp_path: Path) -> None:
    left = _snapshot(tmp_path / "left", [_line("100"), _line("200"), _line("300")], month=5)
    right = _snapshot(tmp_path / "right", [_line("100"), _line("200"), _line("300")], month=6)
    relations = infer_relations([left, right])
    groups = resolve_groups([left, right], relations, [], Pseudonymizer(SEED))
    assert _blocking_conflicts(groups[0], strict=True) == ["PERIOD_CONFLICT"]

    declaration = ManifestGroup(
        "equivalent_002",
        RelationType.EXACT_EQUIVALENT,
        (left.source.resolve(), right.source.resolve()),
        "entity_002",
        left.source.resolve(),
    )
    confirmed = resolve_groups([left, right], relations, [declaration], Pseudonymizer(SEED))[0]
    assert _blocking_conflicts(confirmed, strict=True) == []


def test_atomic_group_failure_promotes_no_member(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    xlsx = tmp_path / "one.xlsx"
    xml = tmp_path / "two.xml"
    _make_xlsx_b(xlsx)
    _make_xml(xml)
    output = tmp_path / "out"
    manifest = [
        ManifestGroup(
            "atomic_001",
            RelationType.EXACT_EQUIVALENT,
            (xlsx.resolve(), xml.resolve()),
            "entity_003",
            xlsx.resolve(),
        )
    ]
    with PseudonymRegistry(output / "registry.sqlite") as registry:
        processor = BatchProcessor(SEED, registry=registry)

        def fail_apply(*args: object, **kwargs: object) -> object:
            raise AdapterError("TEST_MEMBER_FAILURE")

        monkeypatch.setattr(processor.adapters[".xml"], "apply", fail_apply)
        run = processor.run([xlsx.resolve(), xml.resolve()], output, manifest=manifest)

    assert len(run.results) == 2 and not any(item.success for item in run.results)
    assert {item.atomic_state for item in run.results} == {"GROUP_FAILED"}
    assert not list(output.glob("anonimizado_*"))


def test_cli_never_uses_a_source_as_report_destination(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    _make_xlsx_b(source)
    original = source.read_bytes()
    monkeypatch.setenv("BALANCE_ANON_SEED", SEED)

    code = main(
        [
            "--input", str(source),
            "--output", str(tmp_path / "out"),
            "--discover",
            "--report", str(source),
        ]
    )

    assert code == 2
    assert source.read_bytes() == original


def test_cli_retry_report_processes_only_previously_failed_files(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    successful = tmp_path / "successful.xlsx"
    failed = tmp_path / "failed.xml"
    _make_xlsx_b(successful)
    _make_xml(failed)
    pseudo = Pseudonymizer(SEED)
    retry_report = tmp_path / "previous-report.json"
    retry_report.write_text(
        json.dumps(
            {
                "version": 3,
                "archivos": [
                    {
                        "id_archivo": pseudo.token("report-file", str(successful.resolve()), 32),
                        "exitoso": True,
                    },
                    {
                        "id_archivo": pseudo.token("report-file", str(failed.resolve()), 32),
                        "exitoso": False,
                    },
                ],
            }
        ),
        encoding="utf-8",
    )
    report = tmp_path / "retry-result.json"
    monkeypatch.setenv("BALANCE_ANON_SEED", SEED)

    code = main(
        [
            "--input", str(tmp_path),
            "--output", str(tmp_path / "out"),
            "--discover",
            "--retry-report", str(retry_report),
            "--report", str(report),
        ]
    )

    assert code == 0
    payload = json.loads(report.read_text(encoding="utf-8"))
    assert payload["resumen"]["archivos"] == 1
    assert payload["archivos"][0]["id_archivo"] == pseudo.token(
        "report-file", str(failed.resolve()), 32
    )


def test_discover_generates_reviewable_manifest_proposal(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    small = tmp_path / "small.xml"
    large = tmp_path / "large.xml"
    shared_rows = (
        ("101.01", "1000.00", "200.00", "50.00", "1150.00"),
        ("102.01", "10.00", "3.00", "1.00", "12.00"),
        ("103.01", "20.00", "4.00", "2.00", "22.00"),
    )
    _make_xml(small, rows=shared_rows)
    _make_xml(
        large,
        rows=shared_rows + (("104.01", "30.00", "5.00", "3.00", "32.00"),),
    )
    proposal = tmp_path / "equivalence.proposed.yml"
    monkeypatch.setenv("BALANCE_ANON_SEED", SEED)

    code = main(
        [
            "--input", str(tmp_path),
            "--output", str(tmp_path / "out"),
            "--discover",
            "--manifest-proposal", str(proposal),
            "--report", str(tmp_path / "discover.json"),
        ]
    )

    assert code == 0
    payload = yaml.safe_load(proposal.read_text(encoding="utf-8"))
    assert len(payload["groups"]) == 1
    group = payload["groups"][0]
    assert group["relation"] == "projection"
    assert group["entity_id"].startswith("entity_")
    assert set(group["files"]) == {"small.xml", "large.xml"}
    assert "metadata_source" not in group
    resolved = load_manifest(proposal)
    assert len(resolved) == 1


@pytest.mark.private_integration
def test_sanitized_samples_have_expected_content_relations_and_standalone_processing(tmp_path: Path) -> None:
    configured = os.environ.get("BALANCE_SAMPLE_DIR")
    if not configured:
        pytest.skip("Defina BALANCE_SAMPLE_DIR para ejecutar las muestras sanitizadas locales.")
    sources = list_input_files(Path(configured))
    discovery = BatchProcessor(SEED, registry=None).run(
        sources, tmp_path / "discover", discover_only=True
    )
    assert len(discovery.snapshots) == 12

    projection = next(
        item for item in discovery.relations
        if item.relation == RelationType.PROJECTION and item.shared_accounts == 33
    )
    left = next(item for item in discovery.snapshots if item.source == projection.left)
    right = next(item for item in discovery.snapshots if item.source == projection.right)
    assert sorted((len(left.ledger_lines), len(right.ledger_lines))) == [34, 393]

    left_map, right_map = _comparison_maps(left, right)
    shared = set(left_map) & set(right_map)
    assert len(shared) == 33
    for key in shared:
        assert tuple(left_map[key].amounts[name] for name in ("saldo_inicial", "debe", "haber", "saldo_final")) == tuple(
            right_map[key].amounts[name] for name in ("saldo_inicial", "debe", "haber", "saldo_final")
        )

    exact = next(
        item for item in discovery.relations
        if item.relation == RelationType.EXACT_EQUIVALENT and item.shared_accounts == 61
    )
    assert {"PERIOD_CONFLICT", "PRINT_DATE_CONFLICT"} <= set(exact.conflicts)
    assert {
        next(item for item in discovery.snapshots if item.source == exact.left).adapter,
        next(item for item in discovery.snapshots if item.source == exact.right).adapter,
    } == {"pdf", "xlsx"}

    series = next(item for item in discovery.relations if item.relation == RelationType.SERIES)
    assert series.relation != RelationType.EXACT_EQUIVALENT
    assert {
        next(item for item in discovery.snapshots if item.source == series.left).adapter,
        next(item for item in discovery.snapshots if item.source == series.right).adapter,
    } == {"pdf", "xml"}

    xml_counts = sorted(
        len(item.ledger_lines) for item in discovery.snapshots if item.adapter == "xml"
    )
    assert xml_counts == [393, 450]
    xlsx_a = next(item for item in discovery.snapshots if item.profile == "XLSX_CONTPAQ_8_COLUMNS")
    associated = [
        span for span in xlsx_a.sensitive_spans
        if span.category in {Category.ASSOCIATED_BANK, Category.ASSOCIATED_ENTITY}
    ]
    assert len(associated) == 2
    xlsx_c = next(item for item in discovery.snapshots if item.profile == "XLSX_MERGED_HEADER")
    assert not any(line.sensitive_spans for line in xlsx_c.ledger_lines)
    assert sum(
        span.category == Category.RASTER_IMAGE for span in xlsx_c.sensitive_spans
    ) == 1

    standalone = _process([xlsx_a.source], tmp_path / "standalone")
    assert standalone.results[0].success
