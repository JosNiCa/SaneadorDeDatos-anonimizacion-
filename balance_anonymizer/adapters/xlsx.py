"""Adaptador XLSX conservador basado en etiquetas y estructura."""

from __future__ import annotations

import copy
import html
import io
import os
import posixpath
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.writer.excel import ExcelWriter
from lxml import etree

from ..models import (
    AnonymizationPlan,
    Category,
    DocumentSnapshot,
    FormatLocation,
    LedgerLine,
    OwnerIdentity,
    SensitiveSpan,
)
from ..pseudonyms import Pseudonymizer, normalize
from ..relations import normalize_account_code
from .base import AdapterError, AdapterOutput
from .common import (
    DATE_RE,
    RFC_RE,
    YEAR_MONTH_RE,
    decimal_value,
    detect_description_span,
    format_date_like,
    parse_temporal_text,
    replacement_for_description,
)


XLSX_A = "XLSX_CONTPAQ_8_COLUMNS"
XLSX_B = "XLSX_COMBINED_ACCOUNT"
XLSX_C = "XLSX_MERGED_HEADER"
ACCOUNT_RE = re.compile(r"^\d+(?:[.-]\d+)*$")
COMBINED_ACCOUNT_RE = re.compile(r"^(?P<code>\d+(?:[.-]\d+)*)\s+(?P<description>.+)$")
OWNER_EXCLUSIONS = (
    "BALANZA", "PERIODO", "FECHA", "MONEDA", "SALDO", "CUENTA",
    "DEBE", "HABER", "CARGO", "ABONO", "HOJA", "CONTPAQ", "RFC",
    "DIRECCION", "DOMICILIO", "POBLACION", "CEDULA",
    "NATURALEZA", "NOMBRE", "DESCRIPCION", "DEUDOR", "ACREEDOR",
    "INICIALES", "ACTUALES",
)
UNSUPPORTED_PARTS = (
    "vbaproject", "activex", "embeddings/", "oleobject", "controls/",
)


_XLSX_DISCOVERY_CODES = {
    "El libro excede el límite de tamaño permitido.": "XLSX_ARCHIVE_TOO_LARGE",
    "El XLSX no es un contenedor OOXML válido.": "XLSX_INVALID_CONTAINER",
    "El libro excede los límites estructurales permitidos.": "XLSX_STRUCTURE_LIMIT",
    "No se reconoció una familia XLSX compatible.": "XLSX_PROFILE_UNRECOGNIZED",
    "No se localizaron renglones contables en el XLSX.": "XLSX_LEDGER_UNREADABLE",
    "No se localizó el propietario del XLSX.": "XLSX_OWNER_UNREADABLE",
}


def _xlsx_discovery_code(error: BaseException) -> str:
    """Devuelve un código que no incorpora contenido de la hoja."""
    return _XLSX_DISCOVERY_CODES.get(str(error), "XLSX_DISCOVERY_FAILED")
MAX_ARCHIVE_SIZE = 100 * 1024 * 1024
MAX_UNCOMPRESSED_SIZE = 250 * 1024 * 1024
MAX_PARTS = 5000

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _worksheet_part_paths(parts: dict[str, bytes]) -> dict[str, str]:
    """Relaciona nombres de hoja con sus partes OOXML sin asumir sheet1.xml."""

    workbook = etree.fromstring(parts["xl/workbook.xml"])
    relationships = etree.fromstring(parts["xl/_rels/workbook.xml.rels"])
    targets = {
        item.get("Id"): item.get("Target")
        for item in relationships.findall(f"{{{_PACKAGE_REL_NS}}}Relationship")
    }
    result: dict[str, str] = {}
    for sheet in workbook.findall(f".//{{{_SPREADSHEET_NS}}}sheet"):
        relationship_id = sheet.get(f"{{{_OFFICE_REL_NS}}}id")
        target = targets.get(relationship_id)
        name = sheet.get("name")
        if not target or not name:
            continue
        if target.startswith("/"):
            part = target.lstrip("/")
        else:
            part = posixpath.normpath(posixpath.join("xl", target))
        result[name] = part
    return result


def _shared_string_as_inline(cell: Any, shared_strings: list[Any]) -> Any:
    """Convierte una referencia compartida en texto inline sin copiar PII huérfana."""

    if cell.get("t") != "s":
        return cell
    value = cell.find(f"{{{_SPREADSHEET_NS}}}v")
    if value is None or value.text is None:
        return cell
    try:
        shared = shared_strings[int(value.text)]
    except (IndexError, ValueError):
        raise AdapterError("El XLSX contiene una referencia sharedStrings inválida.")
    replacement = copy.deepcopy(cell)
    replacement.set("t", "inlineStr")
    for child in list(replacement):
        replacement.remove(child)
    inline = etree.Element(f"{{{_SPREADSHEET_NS}}}is")
    for child in shared:
        inline.append(copy.deepcopy(child))
    replacement.append(inline)
    return replacement


def _cell_column_index(coordinate: str) -> int:
    letters = re.match(r"[A-Z]+", coordinate)
    if letters is None:
        raise AdapterError("Coordenada OOXML de celda inválida.")
    value = 0
    for character in letters.group(0):
        value = value * 26 + ord(character) - ord("A") + 1
    return value


def _style_index_mapping(source_parts: dict[str, bytes], target_parts: dict[str, bytes]) -> dict[str, str | None]:
    """Mapea estilos equivalentes entre el catálogo de origen y el normalizado."""

    source_style = source_parts.get("xl/styles.xml")
    target_style = target_parts.get("xl/styles.xml")
    if source_style is None or target_style is None:
        return {}
    source_root = etree.fromstring(source_style)
    target_root = etree.fromstring(target_style)
    source_xfs = source_root.findall(f".//{{{_SPREADSHEET_NS}}}cellXfs/{{{_SPREADSHEET_NS}}}xf")
    target_xfs = target_root.findall(f".//{{{_SPREADSHEET_NS}}}cellXfs/{{{_SPREADSHEET_NS}}}xf")
    target_by_xml: dict[bytes, str] = {
        etree.tostring(item, method="c14n"): str(index)
        for index, item in enumerate(target_xfs)
    }
    return {
        str(index): target_by_xml[etree.tostring(item, method="c14n")]
        for index, item in enumerate(source_xfs)
        if etree.tostring(item, method="c14n") in target_by_xml
    }


def _restore_non_target_cell_payloads(
    source: Path,
    target: Path,
    target_cells: set[tuple[str, str]],
) -> set[tuple[str, str]]:
    """Restaura celdas no objetivo que el escritor de openpyxl normaliza.

    Algunos productores guardan cadenas vacías y decimales IEEE con una
    representación que openpyxl vuelve a serializar. Se restaura cada nodo de
    celda no sensible desde el paquete fuente. Las cadenas compartidas se
    convierten a inline para no reincorporar una tabla completa que pudiera
    contener valores sensibles ya sustituidos.
    """

    ignored_empty_cells: set[tuple[str, str]] = set()
    with zipfile.ZipFile(source) as archive:
        source_parts = {name: archive.read(name) for name in archive.namelist()}
    with zipfile.ZipFile(target) as archive:
        target_infos = archive.infolist()
        target_parts = {item.filename: archive.read(item.filename) for item in target_infos}

    source_sheets = _worksheet_part_paths(source_parts)
    target_sheets = _worksheet_part_paths(target_parts)
    shared_root = (
        etree.fromstring(source_parts["xl/sharedStrings.xml"])
        if "xl/sharedStrings.xml" in source_parts
        else None
    )
    shared_strings = (
        shared_root.findall(f"{{{_SPREADSHEET_NS}}}si") if shared_root is not None else []
    )
    normalized_style_ids = _style_index_mapping(source_parts, target_parts)

    for sheet_name, source_part in source_sheets.items():
        target_part = target_sheets.get(sheet_name)
        if target_part is None or source_part not in source_parts or target_part not in target_parts:
            raise AdapterError("Cambió la relación OOXML de una hoja XLSX.")
        source_root = etree.fromstring(source_parts[source_part])
        target_root = etree.fromstring(target_parts[target_part])
        source_cells = {
            cell.get("r"): cell
            for cell in source_root.findall(f".//{{{_SPREADSHEET_NS}}}c")
            if cell.get("r")
        }
        target_nodes = {
            cell.get("r"): cell
            for cell in target_root.findall(f".//{{{_SPREADSHEET_NS}}}c")
            if cell.get("r")
        }
        target_rows = {
            row.get("r"): row
            for row in target_root.findall(f".//{{{_SPREADSHEET_NS}}}row")
            if row.get("r")
        }
        style_ids: dict[str, str | None] = dict(normalized_style_ids)
        for coordinate, source_cell in source_cells.items():
            target_cell = target_nodes.get(coordinate)
            source_style = source_cell.get("s")
            if target_cell is not None and source_style is not None:
                style_ids.setdefault(source_style, target_cell.get("s"))
        for coordinate, target_cell in list(target_nodes.items()):
            if (sheet_name, coordinate) in target_cells:
                continue
            source_cell = source_cells.get(coordinate)
            if source_cell is None:
                target_cell.getparent().remove(target_cell)
                continue
            replacement = _shared_string_as_inline(copy.deepcopy(source_cell), shared_strings)
            target_style = target_cell.get("s")
            if target_style is None:
                replacement.attrib.pop("s", None)
            else:
                replacement.set("s", target_style)
            target_cell.getparent().replace(target_cell, replacement)

        for coordinate, source_cell in source_cells.items():
            if coordinate in target_nodes or (sheet_name, coordinate) in target_cells:
                continue
            if not list(source_cell):
                # Celdas vacías con estilos heredados no materiales pueden
                # apuntar a xfs que el escritor ya normalizó. No contienen
                # valor, fórmula ni texto y recrearlas volvería inválido el
                # paquete para lectores estrictos como Numbers.
                ignored_empty_cells.add((sheet_name, coordinate))
                continue
            row_number = re.sub(r"\D", "", coordinate)
            row = target_rows.get(row_number)
            if row is None:
                raise AdapterError("Desapareció un renglón OOXML del XLSX.")
            replacement = _shared_string_as_inline(copy.deepcopy(source_cell), shared_strings)
            source_style = source_cell.get("s")
            if source_style is not None:
                if source_style not in style_ids:
                    raise AdapterError("No se pudo preservar el estilo de una celda XLSX.")
                mapped_style = style_ids[source_style]
                if mapped_style is None:
                    replacement.attrib.pop("s", None)
                else:
                    replacement.set("s", mapped_style)
            cells = [
                item for item in row.findall(f"{{{_SPREADSHEET_NS}}}c")
                if item.get("r")
            ]
            insertion = next(
                (
                    index for index, item in enumerate(cells)
                    if _cell_column_index(item.get("r") or "") > _cell_column_index(coordinate)
                ),
                len(cells),
            )
            row.insert(insertion, replacement)

        had_declaration = target_parts[target_part].lstrip().startswith(b"<?xml")
        target_parts[target_part] = etree.tostring(
            target_root,
            encoding="UTF-8",
            xml_declaration=had_declaration,
        )

    rewritten = target.with_name(f".{target.name}.rewrite")
    try:
        with zipfile.ZipFile(rewritten, "w") as archive:
            for info in target_infos:
                archive.writestr(info, target_parts[info.filename])
        os.replace(rewritten, target)
    finally:
        rewritten.unlink(missing_ok=True)
    return ignored_empty_cells


def _save_workbook_preserving_properties(book: Any, target: Path) -> None:
    """Guarda sin reemplazar `modified` por la hora real de ejecución.

    `Workbook.save()` asigna automáticamente la hora actual justo antes de
    escribir. La propiedad ya fue seudonimizada por el plan y debe conservarse
    de forma determinista, por lo que se usa el mismo escritor OOXML de
    openpyxl omitiendo únicamente esa asignación automática.
    """

    archive = zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED, allowZip64=True)
    try:
        ExcelWriter(book, archive).save()
    except Exception:
        archive.close()
        raise


def _load_workbook_compatible(source: Path) -> Any:
    """Abre OOXML tolerado por Excel pero rechazado por openpyxl.

    Algunos generadores colocan ``customWidth`` en ``sheetFormatPr``. Se quita
    únicamente de una copia en memoria para permitir el modelo normal de
    openpyxl; la parte original se restaura al generar la salida.
    """

    compatibility_error: TypeError | None = None
    try:
        return load_workbook(source, data_only=False, keep_links=True, read_only=False)
    except TypeError as exc:
        if "SheetFormatProperties" not in str(exc) or "customWidth" not in str(exc):
            raise
        compatibility_error = exc
    with zipfile.ZipFile(source) as archive:
        infos = archive.infolist()
        parts = {item.filename: archive.read(item.filename) for item in infos}
    changed = False
    for name, payload in list(parts.items()):
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        root = etree.fromstring(payload)
        sheet_format = root.find(f"{{{_SPREADSHEET_NS}}}sheetFormatPr")
        if sheet_format is None or "customWidth" not in sheet_format.attrib:
            continue
        sheet_format.attrib.pop("customWidth")
        parts[name] = etree.tostring(
            root,
            encoding="UTF-8",
            xml_declaration=payload.lstrip().startswith(b"<?xml"),
        )
        changed = True
    if not changed:
        raise compatibility_error or TypeError("No se pudo aplicar la compatibilidad XLSX.")
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        for info in infos:
            archive.writestr(info, parts[info.filename])
    buffer.seek(0)
    book = load_workbook(buffer, data_only=False, keep_links=True, read_only=False)
    book._balance_compatibility_buffer = buffer
    return book


@dataclass(frozen=True)
class SheetProfile:
    name: str
    sheet: str
    header_row: int
    account_column: int
    description_column: int | None
    nature_column: int | None
    monetary_columns: tuple[int, ...]


def _compact_label(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", normalize(str(value or "")))


def _archive_parts(path: Path, *, strict: bool) -> tuple[list[str], list[str]]:
    if path.stat().st_size > MAX_ARCHIVE_SIZE:
        raise AdapterError("El libro excede el límite de tamaño permitido.")
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
    except (OSError, zipfile.BadZipFile) as exc:
        raise AdapterError("El XLSX no es un contenedor OOXML válido.") from exc
    if len(infos) > MAX_PARTS or sum(item.file_size for item in infos) > MAX_UNCOMPRESSED_SIZE:
        raise AdapterError("El libro excede los límites estructurales permitidos.")
    names = [item.filename for item in infos]
    unsupported = [
        name for name in names
        if any(marker in name.lower() for marker in UNSUPPORTED_PARTS)
    ]
    if strict and unsupported:
        raise AdapterError("UNSUPPORTED_XLSX_OBJECT")
    return names, unsupported


def _row_labels(sheet: Any, row: int) -> dict[int, str]:
    return {
        cell.column: _compact_label(cell.value)
        for cell in sheet[row]
        if not isinstance(cell, MergedCell) and cell.value is not None
    }


def _find_profile(book: Any) -> SheetProfile:
    for sheet in book.worksheets:
        for row in range(1, min(sheet.max_row, 50) + 1):
            current = _row_labels(sheet, row)
            labels = set(current.values())
            if {"NOCUENTA", "NATURALEZA", "CUENTA", "SALDOINICIAL", "DEBE", "HABER", "SALDOFINAL"} <= labels:
                by_label = {value: column for column, value in current.items()}
                return SheetProfile(
                    XLSX_C,
                    sheet.title,
                    row,
                    by_label["NOCUENTA"],
                    by_label["CUENTA"],
                    by_label["NATURALEZA"],
                    tuple(by_label[label] for label in ("SALDOINICIAL", "DEBE", "HABER", "SALDOFINAL")),
                )
            monetary_labels = {"SALDOINICIAL", "DEBE", "HABER", "SALDOFINAL"}
            combined_account_column = next(
                (
                    column
                    for column, label in current.items()
                    if label == "CUENTA" or ("CUENTA" in label and "DESCRIPCION" in label)
                ),
                None,
            )
            if monetary_labels <= labels and combined_account_column is not None:
                by_label = {value: column for column, value in current.items()}
                return SheetProfile(
                    XLSX_B,
                    sheet.title,
                    row,
                    combined_account_column,
                    None,
                    None,
                    tuple(by_label[label] for label in ("SALDOINICIAL", "DEBE", "HABER", "SALDOFINAL")),
                )
            next_labels = _row_labels(sheet, row + 1) if row < sheet.max_row else {}
            combined = set(current.values()) | set(next_labels.values())
            if "CARGOS" in combined and "ABONOS" in combined and {"DEUDOR", "ACREEDOR"} <= combined:
                account = next((column for column, value in current.items() if value == "CUENTA"), None)
                description = next((column for column, value in current.items() if value in {"NOMBRE", "DESCRIPCION"}), None)
                if account is None or description is None:
                    continue
                monetary = tuple(sorted(column for column in range(description + 1, sheet.max_column + 1)))
                if len(monetary) != 6:
                    continue
                return SheetProfile(
                    XLSX_A,
                    sheet.title,
                    row,
                    account,
                    description,
                    None,
                    monetary,
                )
    raise AdapterError("No se reconoció una familia XLSX compatible.")


def _cell_location(sheet: Any, cell: Any) -> FormatLocation:
    return FormatLocation("xlsx_cell", sheet=sheet.title, cell=cell.coordinate)


def _ledger_lines(sheet: Any, profile: SheetProfile) -> list[LedgerLine]:
    result: list[LedgerLine] = []
    for row in range(profile.header_row + 1, sheet.max_row + 1):
        account_cell = sheet.cell(row, profile.account_column)
        account_value = account_cell.value
        if not isinstance(account_value, str):
            continue
        account_text = account_value.strip()
        description = ""
        if profile.name == XLSX_B:
            match = COMBINED_ACCOUNT_RE.fullmatch(account_text)
            if not match:
                continue
            account_text, description = match.group("code"), match.group("description")
        else:
            if not ACCOUNT_RE.fullmatch(account_text):
                continue
            value = sheet.cell(row, profile.description_column or 0).value
            description = str(value) if value is not None else ""
        nature_value = sheet.cell(row, profile.nature_column).value if profile.nature_column else None
        nature = str(nature_value).strip() if nature_value is not None else None
        raw_values: list[Decimal] = []
        representations: list[str] = []
        for column in profile.monetary_columns:
            value = sheet.cell(row, column).value
            parsed, representation = decimal_value(value if value is not None else 0)
            raw_values.append(parsed)
            representations.append(representation)
        if profile.name == XLSX_A:
            initial_debit, initial_credit, debit, credit, final_debit, final_credit = raw_values
            amounts = {
                "saldo_inicial": initial_debit - initial_credit,
                "debe": debit,
                "haber": credit,
                "saldo_final": final_debit - final_credit,
                "saldo_inicial_deudor": initial_debit,
                "saldo_inicial_acreedor": initial_credit,
                "saldo_final_deudor": final_debit,
                "saldo_final_acreedor": final_credit,
            }
            amount_representations = {
                key: value
                for key, value in zip(
                    (
                        "saldo_inicial_deudor", "saldo_inicial_acreedor", "debe", "haber",
                        "saldo_final_deudor", "saldo_final_acreedor",
                    ),
                    representations,
                )
            }
        else:
            amounts = dict(zip(("saldo_inicial", "debe", "haber", "saldo_final"), raw_values))
            amount_representations = dict(
                zip(("saldo_inicial", "debe", "haber", "saldo_final"), representations)
            )
        location = _cell_location(sheet, account_cell)
        description_location = (
            location
            if profile.name == XLSX_B
            else _cell_location(sheet, sheet.cell(row, profile.description_column or 0))
        )
        sensitive = detect_description_span(description, description_location)
        result.append(
            LedgerLine(
                account_text,
                normalize_account_code(account_text),
                nature,
                description,
                amounts,
                amount_representations,
                location,
                [sensitive] if sensitive else [],
            )
        )
    if not result:
        raise AdapterError("No se localizaron renglones contables en el XLSX.")
    return result


def _header_cells(book: Any, profile: SheetProfile) -> list[tuple[Any, Any]]:
    result: list[tuple[Any, Any]] = []
    for sheet in book.worksheets:
        limit = min(sheet.max_row, 50)
        if sheet.title == profile.sheet:
            for row_number in range(profile.header_row + 1, sheet.max_row + 1):
                value = sheet.cell(row_number, profile.account_column).value
                text = str(value).strip() if isinstance(value, str) else ""
                matches = (
                    COMBINED_ACCOUNT_RE.fullmatch(text)
                    if profile.name == XLSX_B
                    else ACCOUNT_RE.fullmatch(text)
                )
                if matches:
                    limit = row_number - 1
                    break
        for row in sheet.iter_rows(min_row=1, max_row=max(1, limit)):
            for cell in row:
                if not isinstance(cell, MergedCell) and isinstance(cell.value, str) and cell.value.strip():
                    result.append((sheet, cell))
    return result


def _owner_and_spans(
    book: Any,
    profile: SheetProfile,
    ledger: list[LedgerLine],
) -> tuple[OwnerIdentity, list[SensitiveSpan]]:
    headers = _header_cells(book, profile)
    rfc_match: tuple[Any, Any, re.Match[str]] | None = None
    for sheet, cell in headers:
        if match := RFC_RE.search(cell.value):
            rfc_match = sheet, cell, match
            break
    owner_name: str | None = None
    owner_cell: tuple[Any, Any] | None = None
    owner_label = re.compile(r"^\s*(?:NOMBRE|RAZ[ÓO]N\s+SOCIAL)\s*:?\s*(.*)$", re.I)
    for sheet, cell in headers:
        match = owner_label.match(cell.value)
        if not match:
            continue
        candidate = match.group(1).strip()
        if not candidate:
            adjacent = sheet.cell(cell.row, cell.column + 1)
            candidate = str(adjacent.value).strip() if adjacent.value is not None else ""
            if candidate:
                owner_cell = sheet, adjacent
        else:
            owner_cell = sheet, cell
        if candidate:
            candidate = RFC_RE.sub("", candidate).strip(" :-;,.\t")
            if candidate:
                owner_name = candidate
                break
    if rfc_match:
        sheet, cell, match = rfc_match
        candidate = (cell.value[: match.start()] + " " + cell.value[match.end() :]).strip(" :-;,.")
        if owner_name is None and candidate and not any(word in normalize(candidate) for word in OWNER_EXCLUSIONS):
            owner_name, owner_cell = candidate, (sheet, cell)
    if owner_name is None:
        candidates: list[tuple[int, str, Any, Any]] = []
        for sheet, cell in headers:
            text = cell.value.strip()
            normalized = normalize(text)
            if sum(char.isalpha() for char in text) < 4:
                continue
            if RFC_RE.search(text) or any(word in normalized for word in OWNER_EXCLUSIONS):
                continue
            candidates.append((sum(char.isalpha() for char in text), text, sheet, cell))
        if candidates:
            _, owner_name, sheet, cell = max(
                candidates,
                key=lambda item: (
                    item[2].title == profile.sheet,
                    item[0],
                    normalize(item[1]),
                    item[2].title,
                    item[3].coordinate,
                ),
            )
            owner_cell = sheet, cell
    if owner_name is None and rfc_match is None:
        raise AdapterError("No se localizó el propietario del XLSX.")

    owner = OwnerIdentity(name=owner_name, rfc=rfc_match[2].group(1) if rfc_match else None)
    spans: list[SensitiveSpan] = []
    if owner_name and owner_cell:
        location = _cell_location(*owner_cell)
        owner.locations["name"] = location
        spans.append(SensitiveSpan(Category.COMPANY, owner_name, location, confidence=0.99))
    if rfc_match:
        sheet, cell, match = rfc_match
        location = _cell_location(sheet, cell)
        owner.locations["rfc"] = location
        spans.append(SensitiveSpan(Category.RFC, match.group(1), location, confidence=1.0))

    labeled_fields = (
        ("address", Category.ADDRESS, re.compile(r"^\s*(?:DIRECCI[ÓO]N|DOMICILIO)\s*:?\s*(.*)$", re.I)),
        ("population", Category.POPULATION, re.compile(r"^\s*POBLACI[ÓO]N\s*:?\s*(.*)$", re.I)),
        ("certificate", Category.CERTIFICATE, re.compile(r"^\s*C[ÉE]DULA\s*:?\s*(.*)$", re.I)),
    )
    for attribute, category, pattern in labeled_fields:
        for sheet, cell in headers:
            match = pattern.match(cell.value)
            if not match:
                continue
            original = match.group(1).strip()
            target_cell = cell
            if not original:
                adjacent = sheet.cell(cell.row, cell.column + 1)
                original = str(adjacent.value).strip() if adjacent.value is not None else ""
                target_cell = adjacent
            if not original:
                continue
            setattr(owner, attribute, original)
            location = _cell_location(sheet, target_cell)
            owner.locations[attribute] = location
            spans.append(SensitiveSpan(category, original, location, confidence=0.99))
            break

    header_temporals = [
        (cell.value, _cell_location(sheet, cell))
        for sheet, cell in headers
        if DATE_RE.search(cell.value)
        or (YEAR_MONTH_RE.search(cell.value) and "PERIODO" in normalize(cell.value))
        or ("EJERCICIO" in normalize(cell.value) and "PERIODO" in normalize(cell.value))
    ]
    for value, location in header_temporals:
        category = Category.PRINT_DATE if "IMPRESION" in normalize(value) or normalize(value).startswith("FECHA:") else Category.HEADER_DATE
        spans.append(SensitiveSpan(category, value, location, confidence=0.99))

    for sheet, cell in headers:
        if "CONTPAQ" in normalize(cell.value):
            spans.append(
                SensitiveSpan(Category.TEXT_LOGO, cell.value, _cell_location(sheet, cell), confidence=1.0)
            )

    for line in ledger:
        spans.extend(line.sensitive_spans)

    # Repeticiones visibles u ocultas del propietario y RFC.
    originals = [
        (Category.COMPANY, owner.name),
        (Category.RFC, owner.rfc),
        (Category.ADDRESS, owner.address),
        (Category.POPULATION, owner.population),
        (Category.CERTIFICATE, owner.certificate),
    ]
    cell_originals = originals + [
        (span.category, span.original)
        for span in spans
        if span.location.kind == "xlsx_cell"
        and span.category in {
            Category.HEADER_DATE,
            Category.PRINT_DATE,
            Category.PERIOD_RANGE,
            Category.EXERCISE_PERIOD,
            Category.TEXT_LOGO,
        }
    ]
    existing = {(span.category, span.location.sheet, span.location.cell, span.original) for span in spans}
    temporal_cells = {
        (span.location.sheet, span.location.cell)
        for span in spans
        if span.location.kind == "xlsx_cell"
        and span.category in {
            Category.HEADER_DATE,
            Category.PRINT_DATE,
            Category.PERIOD_RANGE,
            Category.EXERCISE_PERIOD,
        }
    }
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if isinstance(cell.value, str):
                    for category, original in cell_originals:
                        if original and original in cell.value:
                            key = (category, sheet.title, cell.coordinate, original)
                            temporal_category = category in {
                                Category.HEADER_DATE,
                                Category.PRINT_DATE,
                                Category.PERIOD_RANGE,
                                Category.EXERCISE_PERIOD,
                            }
                            if (
                                key not in existing
                                and not (temporal_category and (sheet.title, cell.coordinate) in temporal_cells)
                            ):
                                spans.append(
                                    SensitiveSpan(category, original, _cell_location(sheet, cell), confidence=1.0)
                                )
                                existing.add(key)
                                if temporal_category:
                                    temporal_cells.add((sheet.title, cell.coordinate))
                if cell.comment:
                    for category, original in originals:
                        if original and original in cell.comment.text:
                            spans.append(
                                SensitiveSpan(
                                    category,
                                    original,
                                    FormatLocation("xlsx_comment", sheet=sheet.title, cell=cell.coordinate),
                                )
                            )
    for attribute in ("title", "subject", "creator", "lastModifiedBy", "description", "keywords"):
        value = getattr(book.properties, attribute, None)
        if not isinstance(value, str):
            continue
        for category, original in originals:
            if original and original in value:
                spans.append(
                    SensitiveSpan(
                        category,
                        original,
                        FormatLocation("xlsx_property", part=attribute),
                    )
                )
    for attribute, category in (("created", Category.CREATION_DATE), ("modified", Category.CREATION_DATE)):
        value = getattr(book.properties, attribute, None)
        if isinstance(value, (date, datetime)):
            spans.append(
                SensitiveSpan(
                    category,
                    value.isoformat(),
                    FormatLocation("xlsx_property_datetime", part=attribute),
                )
            )
    header_footer_names = (
        "oddHeader", "oddFooter", "evenHeader", "evenFooter", "firstHeader", "firstFooter",
    )
    for sheet in book.worksheets:
        for container_name in header_footer_names:
            container = getattr(sheet, container_name)
            for side in ("left", "center", "right"):
                text = getattr(container, side).text
                if not isinstance(text, str) or not text:
                    continue
                location = FormatLocation(
                    "xlsx_header_footer",
                    sheet=sheet.title,
                    part=f"{container_name}.{side}",
                )
                for category, original in originals:
                    if original and original in text:
                        spans.append(SensitiveSpan(category, original, location))
                for match in DATE_RE.finditer(text):
                    spans.append(SensitiveSpan(Category.FOOTER_DATE, match.group(0), location))
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or cell.hyperlink is None:
                    continue
                for field_name in ("target", "location", "tooltip", "display"):
                    value = getattr(cell.hyperlink, field_name, None)
                    if not isinstance(value, str):
                        continue
                    for category, original in originals:
                        if original and original in value:
                            spans.append(
                                SensitiveSpan(
                                    category,
                                    original,
                                    FormatLocation(
                                        "xlsx_hyperlink",
                                        sheet=sheet.title,
                                        cell=cell.coordinate,
                                        part=field_name,
                                    ),
                                )
                            )
    for index, defined_name in enumerate(book.defined_names.values()):
        for field_name in ("attr_text", "comment", "description"):
            value = getattr(defined_name, field_name, None)
            if not isinstance(value, str):
                continue
            for category, original in originals:
                if original and original in value:
                    spans.append(
                        SensitiveSpan(
                            category,
                            original,
                            FormatLocation(
                                "xlsx_defined_name",
                                part=f"{index}:{field_name}",
                            ),
                        )
                    )
    return owner, spans


def _image_spans(book: Any, profile: SheetProfile, *, strict: bool) -> list[SensitiveSpan]:
    result: list[SensitiveSpan] = []
    for sheet in book.worksheets:
        for index, image in enumerate(sheet._images):
            anchor = getattr(image, "anchor", None)
            marker = getattr(anchor, "_from", None)
            row = int(marker.row) + 1 if marker is not None else None
            in_header = row is not None and row <= profile.header_row
            small = float(getattr(image, "width", 10_000)) <= 400 and float(getattr(image, "height", 10_000)) <= 200
            end_marker = getattr(anchor, "to", None)
            compact_anchor = bool(
                marker is not None
                and end_marker is not None
                and 0 <= int(end_marker.row) - int(marker.row) <= 4
                and 0 <= int(end_marker.col) - int(marker.col) <= 4
            )
            if in_header and (small or compact_anchor):
                result.append(
                    SensitiveSpan(
                        Category.RASTER_IMAGE,
                        f"image:{index}",
                        FormatLocation("xlsx_image", sheet=sheet.title, part=str(index)),
                        confidence=0.99,
                    )
                )
    return result


def _dimensions_state(sheet: Any) -> dict[str, Any]:
    return {
        "title": sheet.title,
        "state": sheet.sheet_state,
        "merged": tuple(sorted(str(value) for value in sheet.merged_cells.ranges)),
        "rows": {
            key: (value.height, value.hidden, value.outlineLevel, value.collapsed)
            for key, value in sheet.row_dimensions.items()
        },
        "columns": {
            key: (value.width, value.hidden, value.outline_level, value.collapsed)
            for key, value in sheet.column_dimensions.items()
        },
        "freeze": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "auto_filter": sheet.auto_filter.ref,
        "tables": tuple(sorted(sheet.tables)),
        "images": len(sheet._images),
        "charts": len(sheet._charts),
        "print_area": str(sheet.print_area),
        "print_title_rows": sheet.print_title_rows,
        "print_title_cols": sheet.print_title_cols,
    }


def _cell_state(book: Any) -> dict[tuple[str, str], tuple[Any, ...]]:
    result: dict[tuple[str, str], tuple[Any, ...]] = {}
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                result[(sheet.title, cell.coordinate)] = (
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                    copy.copy(cell.alignment),
                    copy.copy(cell.border),
                    copy.copy(cell.fill),
                    copy.copy(cell.font),
                    copy.copy(cell.protection),
                )
    return result


def _synthetic_owner_value(category: Category, plan: AnonymizationPlan) -> str:
    keys = {
        Category.COMPANY: "name",
        Category.RFC: "rfc",
        Category.ADDRESS: "address",
        Category.POPULATION: "population",
        Category.CERTIFICATE: "certificate",
    }
    key = keys.get(category)
    if key is None or key not in plan.synthetic_owner:
        raise AdapterError("No existe reemplazo compartido para el propietario.")
    return plan.synthetic_owner[key]


def _canonical_date_for(
    category: Category,
    index: int,
    canonical: Any,
) -> date | None:
    if canonical is None:
        return None
    if category == Category.PRINT_DATE:
        return canonical.print_date
    values = [value for value in (canonical.period_start, canonical.period_end) if value]
    if values:
        return values[min(index, len(values) - 1)]
    if canonical.year and canonical.month:
        return date(canonical.year, canonical.month, 1)
    return None


def _replace_temporal_cell(
    original: str,
    category: Category,
    pseudo: Pseudonymizer,
    canonical: Any,
) -> str:
    dates = list(DATE_RE.finditer(original))
    if canonical is not None and dates:
        value = original
        offset = 0
        for index, match in enumerate(dates):
            replacement_date = _canonical_date_for(category, index, canonical)
            if replacement_date is None:
                continue
            formatted = format_date_like(match.group(0), replacement_date)
            start, end = match.start() + offset, match.end() + offset
            value = value[:start] + formatted + value[end:]
            offset += len(formatted) - (match.end() - match.start())
        return value
    year_month = YEAR_MONTH_RE.search(original)
    if year_month:
        if canonical is not None and canonical.year and canonical.month:
            year, month = str(canonical.year), str(canonical.month).zfill(len(year_month.group("month")))
        else:
            year, month = pseudo.exercise_and_period(year_month.group("year"), year_month.group("month"))
        replacement = f"{year}{year_month.group('sep')}{month}"
        return original[: year_month.start()] + replacement + original[year_month.end() :]
    exercise = re.search(r"(EJERCICIO\s*:?\s*)(\d{4})(.*?PER[IÍ]ODO\s*:?\s*)(\d{1,2})", original, re.I)
    if exercise:
        if canonical is not None and canonical.year and canonical.month:
            year, month = str(canonical.year), str(canonical.month).zfill(len(exercise.group(4)))
        else:
            year, month = pseudo.exercise_and_period(exercise.group(2), exercise.group(4))
        return (
            original[: exercise.start()]
            + exercise.group(1) + year + exercise.group(3) + month
            + original[exercise.end() :]
        )
    return pseudo.replace_temporal(original)


def _replace_in_value(value: str, original: str, replacement: str) -> str:
    if original not in value:
        raise AdapterError("El valor sensible ya no coincide con la celda planificada.")
    return value.replace(original, replacement, 1)


def _validate_ledger(original: DocumentSnapshot, generated: DocumentSnapshot) -> None:
    if [line.account_code for line in original.ledger_lines] != [line.account_code for line in generated.ledger_lines]:
        raise AdapterError("La validación XLSX detectó cambios en códigos u orden de cuentas.")
    for before, after in zip(original.ledger_lines, generated.ledger_lines):
        if before.amounts != after.amounts or before.amount_representations != after.amount_representations:
            raise AdapterError("La validación XLSX detectó cambios en importes.")


class XlsxAdapter:
    name = "xlsx"
    suffixes = (".xlsx", ".xlsm")

    def discover(
        self,
        source: Path,
        pseudonymizer: Pseudonymizer,
        *,
        strict: bool,
    ) -> DocumentSnapshot:
        source = source.resolve()
        if source.suffix.lower() == ".xlsm":
            raise AdapterError("UNSUPPORTED_XLSM")
        try:
            parts, unsupported = _archive_parts(source, strict=strict)
        except AdapterError as exc:
            if str(exc) == "UNSUPPORTED_XLSX_OBJECT":
                raise
            raise AdapterError(_xlsx_discovery_code(exc)) from exc
        try:
            book = _load_workbook_compatible(source)
        except Exception as exc:
            raise AdapterError("XLSX_WORKBOOK_UNREADABLE") from exc
        try:
            profile = _find_profile(book)
            sheet = book[profile.sheet]
            ledger = _ledger_lines(sheet, profile)
            owner, spans = _owner_and_spans(book, profile, ledger)
            image_spans = _image_spans(book, profile, strict=strict)
        except AdapterError as exc:
            if str(exc) == "UNSUPPORTED_XLSX_OBJECT":
                raise
            raise AdapterError(_xlsx_discovery_code(exc)) from exc
        spans.extend(image_spans)
        temporal_values = [
            (span.original, span.location)
            for span in spans
            if span.category in {Category.HEADER_DATE, Category.PRINT_DATE, Category.PERIOD_RANGE, Category.EXERCISE_PERIOD}
        ]
        temporal = parse_temporal_text(temporal_values)
        for sheet_item, cell in _header_cells(book, profile):
            if "MONEDA" in normalize(cell.value):
                match = re.search(r"MONEDA\s*:?\s*([A-Z]{3,})", normalize(cell.value))
                if match:
                    temporal.currency = match.group(1)
                    break
        warnings = [f"unsupported_part_count:{len(unsupported)}"] if unsupported else []
        preserved_images = sum(len(item._images) for item in book.worksheets) - len(image_spans)
        if preserved_images:
            warnings.append(f"PRESERVED_NON_LOGO_IMAGES:{preserved_images}")
        structural = {
            "sheet_count": len(book.worksheets),
            "sheets": [_dimensions_state(item) for item in book.worksheets],
            "part_count": len(parts),
            "image_count": sum(len(item._images) for item in book.worksheets),
            "logo_image_count": len(image_spans),
            "formula_count": sum(
                1 for item in book.worksheets for row in item.iter_rows() for cell in row
                if not isinstance(cell, MergedCell) and cell.data_type == "f"
            ),
            "external_link_count": len(book._external_links),
        }
        return DocumentSnapshot(
            source,
            self.name,
            profile.name,
            owner,
            temporal,
            ledger,
            spans,
            warnings,
            structural=structural,
            private={"profile": profile},
        )

    def apply(
        self,
        snapshot: DocumentSnapshot,
        plan: AnonymizationPlan,
        temporary_dir: Path,
        *,
        strict: bool,
    ) -> AdapterOutput:
        try:
            book = _load_workbook_compatible(snapshot.source)
        except Exception as exc:
            raise AdapterError("No se pudo reabrir el XLSX para aplicar el plan.") from exc
        before_cells = _cell_state(book)
        before_sheets = [_dimensions_state(sheet) for sheet in book.worksheets]
        target_cells: set[tuple[str, str]] = set()
        logo_cells: set[tuple[str, str]] = set()
        counts: Counter[str] = Counter()
        grouped: dict[
            tuple[str, str | None, str | None, str | None],
            list[SensitiveSpan],
        ] = defaultdict(list)
        for span in snapshot.sensitive_spans:
            grouped[
                (
                    span.location.kind,
                    span.location.sheet,
                    span.location.cell,
                    span.location.part,
                )
            ].append(span)

        for (kind, sheet_name, cell_locator, part_locator), spans in grouped.items():
            locator = cell_locator or part_locator
            if kind == "xlsx_cell":
                if sheet_name is None or locator is None:
                    raise AdapterError("Ubicación XLSX incompleta.")
                cell = book[sheet_name][locator]
                if not isinstance(cell.value, str):
                    raise AdapterError("Una celda sensible cambió de tipo antes de aplicar el plan.")
                value = cell.value
                for temporal_index, span in enumerate(spans):
                    if span.category in {
                        Category.COMPANY, Category.RFC, Category.ADDRESS,
                        Category.POPULATION, Category.CERTIFICATE,
                    }:
                        replacement = _synthetic_owner_value(span.category, plan)
                        value = _replace_in_value(value, span.original, replacement)
                    elif span.category in {Category.HEADER_DATE, Category.PRINT_DATE, Category.PERIOD_RANGE, Category.EXERCISE_PERIOD}:
                        replacement = _replace_temporal_cell(
                            span.original,
                            span.category,
                            plan.pseudonymizer,
                            plan.canonical_temporal,
                        )
                        value = _replace_in_value(value, span.original, replacement)
                    elif span.category in {Category.ASSOCIATED_ENTITY, Category.ASSOCIATED_BANK}:
                        replacement = replacement_for_description(span, plan.pseudonymizer)
                        value = _replace_in_value(value, span.original, replacement)
                    elif span.category == Category.TEXT_LOGO:
                        value = _replace_in_value(value, span.original, "").strip()
                        logo_cells.add((sheet_name, locator))
                    else:
                        continue
                    counts[span.category.value] += 1
                cell.value = value
                target_cells.add((sheet_name, locator))
            elif kind == "xlsx_comment":
                cell = book[sheet_name][locator]  # type: ignore[index]
                if cell.comment is None:
                    raise AdapterError("El comentario sensible esperado no existe.")
                text = cell.comment.text
                for span in spans:
                    text = _replace_in_value(text, span.original, _synthetic_owner_value(span.category, plan))
                    counts[span.category.value] += 1
                cell.comment.text = text
            elif kind == "xlsx_property":
                value = getattr(book.properties, locator or "", None)
                if not isinstance(value, str):
                    raise AdapterError("La propiedad sensible esperada no existe.")
                for span in spans:
                    value = _replace_in_value(value, span.original, _synthetic_owner_value(span.category, plan))
                    counts[span.category.value] += 1
                setattr(book.properties, locator or "", value)
            elif kind == "xlsx_property_datetime":
                original_value = getattr(book.properties, locator or "", None)
                if not isinstance(original_value, (date, datetime)):
                    raise AdapterError("La fecha de propiedad sensible esperada no existe.")
                shifted = plan.pseudonymizer.shift_date(original_value)
                if isinstance(original_value, datetime):
                    shifted_value: date | datetime = original_value.replace(
                        year=shifted.year,
                        month=shifted.month,
                        day=shifted.day,
                    )
                else:
                    shifted_value = shifted
                setattr(book.properties, locator or "", shifted_value)
                counts[Category.CREATION_DATE.value] += 1
            elif kind == "xlsx_header_footer":
                if sheet_name is None or locator is None:
                    raise AdapterError("Ubicación de encabezado o pie XLSX incompleta.")
                container_name, side = locator.split(".", 1)
                item = getattr(getattr(book[sheet_name], container_name), side)
                value = item.text
                if not isinstance(value, str):
                    raise AdapterError("El encabezado o pie sensible esperado no existe.")
                for span in spans:
                    if span.category in {
                        Category.COMPANY, Category.RFC, Category.ADDRESS,
                        Category.POPULATION, Category.CERTIFICATE,
                    }:
                        replacement = _synthetic_owner_value(span.category, plan)
                    else:
                        replacement = _replace_temporal_cell(
                            span.original,
                            span.category,
                            plan.pseudonymizer,
                            plan.canonical_temporal,
                        )
                    value = _replace_in_value(value, span.original, replacement)
                    counts[span.category.value] += 1
                item.text = value
            elif kind == "xlsx_hyperlink":
                if sheet_name is None or locator is None:
                    raise AdapterError("Ubicación de hipervínculo XLSX incompleta.")
                cell = book[sheet_name][locator]
                if cell.hyperlink is None:
                    raise AdapterError("El hipervínculo sensible esperado no existe.")
                field_name = spans[0].location.part or ""
                value = getattr(cell.hyperlink, field_name, None)
                if not isinstance(value, str):
                    raise AdapterError("El campo sensible del hipervínculo no existe.")
                for span in spans:
                    value = _replace_in_value(
                        value,
                        span.original,
                        _synthetic_owner_value(span.category, plan),
                    )
                    counts[span.category.value] += 1
                setattr(cell.hyperlink, field_name, value)
            elif kind == "xlsx_defined_name":
                if locator is None:
                    raise AdapterError("Ubicación de nombre definido XLSX incompleta.")
                index_text, field_name = locator.split(":", 1)
                names = list(book.defined_names.values())
                try:
                    defined_name = names[int(index_text)]
                except (IndexError, ValueError) as exc:
                    raise AdapterError("El nombre definido sensible esperado no existe.") from exc
                value = getattr(defined_name, field_name, None)
                if not isinstance(value, str):
                    raise AdapterError("El campo sensible del nombre definido no existe.")
                for span in spans:
                    value = _replace_in_value(
                        value,
                        span.original,
                        _synthetic_owner_value(span.category, plan),
                    )
                    counts[span.category.value] += 1
                setattr(defined_name, field_name, value)

        removed_images = 0
        image_indexes: dict[str, set[int]] = defaultdict(set)
        for span in snapshot.sensitive_spans:
            if span.category == Category.RASTER_IMAGE and span.location.sheet and span.location.part:
                image_indexes[span.location.sheet].add(int(span.location.part))
        for sheet_name, indexes in image_indexes.items():
            sheet = book[sheet_name]
            sheet._images = [image for index, image in enumerate(sheet._images) if index not in indexes]
            removed_images += len(indexes)
            counts[Category.RASTER_IMAGE.value] += len(indexes)

        temporary_dir.mkdir(parents=True, exist_ok=True)
        output_source = snapshot.private.get("output_source", snapshot.source)
        target = temporary_dir / (
            f"{plan.output_prefix}anonimizado_"
            f"{plan.pseudonymizer.token('output-file', str(output_source), 16)}.xlsx"
        )
        if target.resolve() == snapshot.source.resolve() or target.exists():
            raise AdapterError("La salida XLSX no puede sobrescribir un archivo existente.")
        try:
            _save_workbook_preserving_properties(book, target)
            ignored_empty_cells = _restore_non_target_cell_payloads(snapshot.source, target, target_cells)
            for key in ignored_empty_cells:
                before_cells.pop(key, None)
        except Exception as exc:
            target.unlink(missing_ok=True)
            raise AdapterError("No se pudo guardar el XLSX temporal.") from exc

        try:
            _archive_parts(target, strict=strict)
            generated_book = _load_workbook_compatible(target)
            after_sheets = [_dimensions_state(sheet) for sheet in generated_book.worksheets]
            after_cells = _cell_state(generated_book)
            if len(before_sheets) != len(after_sheets):
                raise AdapterError("Cambió la cantidad de hojas del XLSX.")
            for before, after in zip(before_sheets, after_sheets):
                expected = dict(before)
                expected["images"] = int(expected["images"]) - len(image_indexes.get(str(before["title"]), set()))
                if expected != after:
                    raise AdapterError("Cambió la estructura no objetivo del XLSX.")
            for key, state in before_cells.items():
                if key not in after_cells:
                    raise AdapterError("Desapareció una celda del XLSX.")
                if key not in target_cells and state != after_cells[key]:
                    raise AdapterError("Cambió una celda no objetivo del XLSX.")
                if key in target_cells:
                    before_format = state[2:] if key in logo_cells else state[1:]
                    after_format = after_cells[key][2:] if key in logo_cells else after_cells[key][1:]
                    if before_format != after_format:
                        changes = tuple(
                            index for index, (left, right) in enumerate(zip(before_format, after_format))
                            if left != right
                        )
                        raise AdapterError(f"Cambió el tipo o estilo de una celda sensible: {key}:{changes}.")
            with zipfile.ZipFile(target) as archive:
                searchable = "\n".join(
                    archive.read(name).decode("utf-8", "ignore")
                    for name in archive.namelist()
                    if name.lower().endswith((".xml", ".rels"))
                )
                for span in snapshot.sensitive_spans:
                    if span.category == Category.RASTER_IMAGE:
                        continue
                    if span.original and (
                        span.original in searchable or html.escape(span.original) in searchable
                    ):
                        raise AdapterError(
                            "Persisten datos sensibles en partes internas del XLSX "
                            f"para la categoría {span.category.value}."
                        )
                media_parts = [name for name in archive.namelist() if name.startswith("xl/media/")]
                if len(media_parts) != snapshot.structural.get("image_count", 0) - removed_images:
                    raise AdapterError("Persisten imágenes de logotipo o relaciones OOXML huérfanas.")
            generated = self.discover(target, plan.pseudonymizer, strict=strict)
            _validate_ledger(snapshot, generated)
        except Exception:
            target.unlink(missing_ok=True)
            raise
        validation = {
            "workbook_reopened": True,
            "non_target_cells_preserved": True,
            "formulas_types_styles_preserved": True,
            "merged_ranges_preserved": True,
            "ooxml_inspected": True,
            "logos_removed": removed_images,
            "ledger_preserved": True,
        }
        return AdapterOutput(
            target,
            snapshot.profile,
            dict(counts),
            validation,
            snapshot.warnings,
            snapshot=generated,
        )
